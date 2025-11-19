import logging
from datetime import datetime
from enum import Enum
from typing import List, Dict, Optional, Any
from importlib import import_module
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import openpyxl  # pyright: ignore[reportMissingImports]
from models import Schedule, ForensicWindow, ComparisonResult, DelayEvent, Activity

try:
    colors = import_module("reportlab.lib.colors")
    pagesizes = import_module("reportlab.lib.pagesizes")
    styles_module = import_module("reportlab.lib.styles")
    units_module = import_module("reportlab.lib.units")
    platypus_module = import_module("reportlab.platypus")
    enums_module = import_module("reportlab.lib.enums")
except ImportError as exc:
    raise ImportError("ReportLab is required for report generation. Install via 'pip install reportlab'.") from exc

letter = pagesizes.letter
A4 = pagesizes.A4
getSampleStyleSheet = styles_module.getSampleStyleSheet
ParagraphStyle = styles_module.ParagraphStyle
inch = units_module.inch
SimpleDocTemplate = platypus_module.SimpleDocTemplate
Paragraph = platypus_module.Paragraph
Spacer = platypus_module.Spacer
Table = platypus_module.Table
TableStyle = platypus_module.TableStyle
PageBreak = platypus_module.PageBreak
Image = platypus_module.Image
KeepTogether = platypus_module.KeepTogether
TA_CENTER = enums_module.TA_CENTER
TA_LEFT = enums_module.TA_LEFT
TA_RIGHT = enums_module.TA_RIGHT

# Configure logging
logger = logging.getLogger(__name__)


class ReportGenerator:
    """
    Generate comprehensive delay analysis reports.
    
    This class handles creating professional PDF and Excel reports
    with charts, tables, and detailed analysis results.
    """
    
    def __init__(self, config: Dict):
        """
        Initialize report generator.
        
        Args:
            config: Configuration dictionary
        """
        self.config = config
        self.report_config = config.get('reports', {})
        self.company_name = self.report_config.get('company_name', 'Company Name')
        logger.info("Report generator initialized")
    
    def generate_pdf_report(self, comparison: ComparisonResult,
                           windows: List[ForensicWindow],
                           output_path: str) -> None:
        """
        Generate comprehensive PDF delay analysis report.
        
        Args:
            comparison: ComparisonResult from schedule comparison
            windows: List of analyzed ForensicWindow objects
            output_path: Path for output PDF file
        """
        logger.info(f"Generating PDF report: {output_path}")
        
        # Create document
        doc = SimpleDocTemplate(
            output_path,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Container for report elements
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Title page
        story.append(Spacer(1, 2*inch))
        story.append(Paragraph("Construction Schedule Delay Analysis Report", title_style))
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph(f"Project: {comparison.baseline_schedule.project_name}", styles['Normal']))
        story.append(Paragraph(f"Report Date: {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
        story.append(Paragraph(f"Prepared by: {self.company_name}", styles['Normal']))
        story.append(PageBreak())
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", heading_style))
        story.extend(self._create_executive_summary(comparison))
        story.append(Spacer(1, 0.3*inch))
        
        # Project Overview
        story.append(Paragraph("Project Overview", heading_style))
        story.extend(self._create_project_overview(comparison))
        story.append(PageBreak())
        
        # Schedule Performance Metrics
        story.append(Paragraph("Schedule Performance Metrics", heading_style))
        story.extend(self._create_performance_metrics(comparison))
        story.append(Spacer(1, 0.3*inch))
        
        # Critical Path Analysis
        story.append(Paragraph("Critical Path Analysis", heading_style))
        story.extend(self._create_critical_path_section(comparison))
        story.append(PageBreak())
        
        # Delayed Activities
        if comparison.delayed_activities:
            story.append(Paragraph("Delayed Activities", heading_style))
            story.extend(self._create_delayed_activities_section(comparison))
            story.append(PageBreak())
        
        # Milestone Analysis
        if comparison.milestone_delays:
            story.append(Paragraph("Milestone Delays", heading_style))
            story.extend(self._create_milestone_section(comparison))
            story.append(Spacer(1, 0.3*inch))
        
        # Forensic Window Analysis
        if windows:
            story.append(Paragraph("Forensic Window Analysis", heading_style))
            story.extend(self._create_window_analysis_section(windows))
            story.append(PageBreak())
        
        # Charts and Visualizations
        story.append(Paragraph("Visual Analysis", heading_style))
        
        # Generate charts
        chart_paths = self._generate_charts(comparison, windows)
        for chart_path in chart_paths:
            try:
                img = Image(chart_path, width=6*inch, height=4*inch)
                story.append(img)
                story.append(Spacer(1, 0.2*inch))
            except Exception as e:
                logger.error(f"Error adding chart {chart_path}: {e}")
        
        # Build PDF
        doc.build(story)
        logger.info("PDF report generation complete")
    
    def _create_executive_summary(self, comparison: ComparisonResult) -> List:
        """Create executive summary section."""
        elements = []
        styles = getSampleStyleSheet()
        
        summary_text = f"""
        This report presents a comprehensive forensic schedule delay analysis comparing the 
        baseline schedule to the current project schedule. The analysis identifies delays, 
        their impacts, and changes to the critical path.
        <br/><br/>
        <b>Key Findings:</b><br/>
        • Overall Project Delay: {comparison.overall_delay} days<br/>
        • Schedule Performance Index (SPI): {comparison.spi:.2f}<br/>
        • Total Delayed Activities: {len(comparison.delayed_activities)}<br/>
        • Activities Became Critical: {len(comparison.new_critical_activities)}<br/>
        • Milestone Delays: {len(comparison.milestone_delays)}<br/>
        """
        
        elements.append(Paragraph(summary_text, styles['Normal']))
        return elements
    
    def _create_project_overview(self, comparison: ComparisonResult) -> List:
        """Create project overview section."""
        elements = []
        styles = getSampleStyleSheet()
        
        baseline = comparison.baseline_schedule
        current = comparison.current_schedule
        
        data = [
            ['Metric', 'Baseline', 'Current', 'Variance'],
            ['Project Start', baseline.start_date.strftime('%Y-%m-%d'), 
             current.start_date.strftime('%Y-%m-%d'), ''],
            ['Project Finish', baseline.finish_date.strftime('%Y-%m-%d'),
             current.finish_date.strftime('%Y-%m-%d'), 
             f"{comparison.overall_delay} days"],
            ['Total Activities', len(baseline.activities),
             len(current.activities), ''],
            ['Critical Activities', len(baseline.get_critical_path()),
             len(current.get_critical_path()), ''],
            ['Data Date', '', current.data_date.strftime('%Y-%m-%d'), '']
        ]
        
        table = Table(data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        return elements
    
    def _create_performance_metrics(self, comparison: ComparisonResult) -> List:
        """Create performance metrics section."""
        elements = []
        styles = getSampleStyleSheet()
        
        spi_status = "Ahead of Schedule" if comparison.spi > 1.0 else "Behind Schedule"
        spi_color = 'green' if comparison.spi > 1.0 else 'red'
        
        metrics_text = f"""
        <b>Schedule Performance Index (SPI):</b> {comparison.spi:.3f} - {spi_status}<br/>
        <b>Completion Variance:</b> {comparison.completion_variance:.1f}%<br/>
        <b>Delayed Activities:</b> {len(comparison.delayed_activities)} 
        ({len(comparison.delayed_activities)/len(comparison.current_schedule.activities)*100:.1f}% of total)<br/>
        <b>Accelerated Activities:</b> {len(comparison.accelerated_activities)}<br/>
        """
        
        elements.append(Paragraph(metrics_text, styles['Normal']))
        return elements
    
    def _create_critical_path_section(self, comparison: ComparisonResult) -> List:
        """Create critical path analysis section."""
        elements = []
        styles = getSampleStyleSheet()
        
        elements.append(Paragraph(
            f"<b>Activities That Became Critical:</b> {len(comparison.new_critical_activities)}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 0.1*inch))
        
        if comparison.new_critical_activities:
            data = [['Activity ID', 'Activity Name', 'Baseline Float', 'Current Float']]
            
            for act in comparison.new_critical_activities[:20]:  # Limit to 20
                baseline_act = comparison.baseline_schedule.activities[act.activity_id]
                data.append([
                    act.activity_id,
                    act.activity_name[:50],
                    f"{baseline_act.total_float:.1f}",
                    f"{act.total_float:.1f}"
                ])
            
            table = Table(data, colWidths=[1*inch, 3*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            
            elements.append(table)
        
        return elements
    
    def _create_delayed_activities_section(self, comparison: ComparisonResult) -> List:
        """Create delayed activities section."""
        elements = []
        styles = getSampleStyleSheet()
        
        # Sort by delay magnitude
        sorted_delays = sorted(
            comparison.delayed_activities,
            key=lambda a: self._get_activity_delay_days(a, comparison.baseline_schedule),
            reverse=True
        )
        
        data = [['Activity ID', 'Activity Name', 'Delay (days)', 'Critical', 'Status']]
        
        for act in sorted_delays[:30]:  # Top 30 delays
            baseline_act = comparison.baseline_schedule.activities[act.activity_id]
            delay_days = self._get_activity_delay_days(act, comparison.baseline_schedule)
            
            data.append([
                act.activity_id,
                act.activity_name[:40],
                f"{delay_days:.1f}",
                "Yes" if act.is_critical else "No",
                act.status.value
            ])
        
        table = Table(data, colWidths=[0.8*inch, 2.5*inch, 0.8*inch, 0.6*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(table)
        return elements
    
    def _create_milestone_section(self, comparison: ComparisonResult) -> List:
        """Create milestone delays section."""
        elements = []
        
        data = [['Milestone', 'Baseline Date', 'Current Date', 'Delay (days)']]
        
        for act_id, delay_days in comparison.milestone_delays.items():
            act = comparison.current_schedule.activities[act_id]
            baseline_act = comparison.baseline_schedule.activities[act_id]
            
            current_date = act.actual_finish or act.finish_date
            
            data.append([
                act.activity_name[:50],
                baseline_act.finish_date.strftime('%Y-%m-%d'),
                current_date.strftime('%Y-%m-%d') if current_date else 'Not Complete',
                f"{delay_days:.1f}"
            ])
        
        table = Table(data, colWidths=[2.5*inch, 1.2*inch, 1.2*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        
        elements.append(table)
        return elements
    
    def _create_window_analysis_section(self, windows: List[ForensicWindow]) -> List:
        """Create forensic window analysis section."""
        elements = []
        styles = getSampleStyleSheet()
        
        data = [['Window', 'Period', 'Total Delays', 'Critical Delays', 'Impact (days)']]
        
        for window in windows:
            data.append([
                window.window_id,
                f"{window.start_date.strftime('%Y-%m-%d')} to {window.end_date.strftime('%Y-%m-%d')}",
                len(window.delays),
                len(window.get_critical_delays()),
                f"{window.get_total_delay():.1f}"
            ])
        
        table = Table(data, colWidths=[1*inch, 2*inch, 0.8*inch, 0.8*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        
        elements.append(table)
        return elements
    
    def _get_activity_delay_days(self, activity: Activity, baseline_schedule: Schedule) -> float:
        """Calculate delay days for an activity."""
        baseline_act = baseline_schedule.activities.get(activity.activity_id)
        if not baseline_act:
            return 0.0
        
        if activity.actual_finish and baseline_act.finish_date:
            return (activity.actual_finish - baseline_act.finish_date).days
        elif activity.finish_date and baseline_act.finish_date:
            return (activity.finish_date - baseline_act.finish_date).days
        
        return 0.0

    def _get_activity_duration(self, activity: Activity) -> Optional[float]:
        """Get best available duration value for an activity."""
        if hasattr(activity, 'duration') and activity.duration is not None:
            return activity.duration
        if getattr(activity, 'original_duration', None) is not None:
            return activity.original_duration
        if getattr(activity, 'remaining_duration', None) is not None:
            return activity.remaining_duration
        return None
    
    def _generate_charts(self, comparison: ComparisonResult,
                        windows: List[ForensicWindow]) -> List[str]:
        """Generate visualization charts."""
        chart_paths = []
        
        try:
            # 1. Delay Histogram
            chart_paths.append(self._create_delay_histogram(comparison))
            
            # 2. Float Changes Chart
            if comparison.float_changes:
                chart_paths.append(self._create_float_changes_chart(comparison))
            
            # 3. Window Analysis Chart
            if windows:
                chart_paths.append(self._create_window_chart(windows))
            
            # 4. Schedule Performance Chart
            chart_paths.append(self._create_spi_chart(comparison))
            
        except Exception as e:
            logger.error(f"Error generating charts: {e}")
        
        return chart_paths
    
    def _create_delay_histogram(self, comparison: ComparisonResult) -> str:
        """Create histogram of delays."""
        fig, ax = plt.subplots(figsize=(10, 6))
        
        delays = [
            self._get_activity_delay_days(act, comparison.baseline_schedule)
            for act in comparison.delayed_activities
        ]
        
        if delays:
            ax.hist(delays, bins=20, color='#1f4788', alpha=0.7, edgecolor='black')
            ax.set_xlabel('Delay (days)', fontsize=12)
            ax.set_ylabel('Number of Activities', fontsize=12)
            ax.set_title('Distribution of Activity Delays', fontsize=14, fontweight='bold')
            ax.grid(True, alpha=0.3)
        
        chart_path = 'temp_delay_histogram.png'
        plt.tight_layout()
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_path
    
    def _create_float_changes_chart(self, comparison: ComparisonResult) -> str:
        """Create chart showing float changes."""
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Get top 20 float changes
        sorted_changes = sorted(
            comparison.float_changes.items(),
            key=lambda x: abs(x[1]),
            reverse=True
        )[:20]
        
        activities = [comparison.current_schedule.activities[act_id].activity_name[:30]
                     for act_id, _ in sorted_changes]
        changes = [change for _, change in sorted_changes]
        
        colors_list = ['red' if c < 0 else 'green' for c in changes]
        
        ax.barh(activities, changes, color=colors_list, alpha=0.7)
        ax.set_xlabel('Float Change (days)', fontsize=12)
        ax.set_title('Top 20 Float Changes', fontsize=14, fontweight='bold')
        ax.axvline(x=0, color='black', linestyle='-', linewidth=0.8)
        ax.grid(True, alpha=0.3, axis='x')
        
        chart_path = 'temp_float_changes.png'
        plt.tight_layout()
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_path
    
    def _create_window_chart(self, windows: List[ForensicWindow]) -> str:
        """Create chart showing delays by window."""
        fig, ax = plt.subplots(figsize=(12, 6))
        
        window_labels = [w.window_id for w in windows]
        total_delays = [len(w.delays) for w in windows]
        critical_delays = [len(w.get_critical_delays()) for w in windows]
        
        x = range(len(window_labels))
        width = 0.35
        
        ax.bar([i - width/2 for i in x], total_delays, width,
               label='Total Delays', color='#1f4788', alpha=0.7)
        ax.bar([i + width/2 for i in x], critical_delays, width,
               label='Critical Delays', color='#d62728', alpha=0.7)
        
        ax.set_xlabel('Window', fontsize=12)
        ax.set_ylabel('Number of Delays', fontsize=12)
        ax.set_title('Delays by Forensic Window', fontsize=14, fontweight='bold')
        ax.set_xticks(x)
        ax.set_xticklabels(window_labels, rotation=45, ha='right')
        ax.legend()
        ax.grid(True, alpha=0.3, axis='y')
        
        chart_path = 'temp_window_chart.png'
        plt.tight_layout()
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_path
    
    def _create_spi_chart(self, comparison: ComparisonResult) -> str:
        """Create schedule performance chart."""
        fig, ax = plt.subplots(figsize=(8, 6))
        
        metrics = ['SPI', 'Target']
        values = [comparison.spi, 1.0]
        colors_list = ['green' if comparison.spi >= 1.0 else 'red', 'blue']
        
        bars = ax.bar(metrics, values, color=colors_list, alpha=0.7)
        ax.set_ylabel('Schedule Performance Index', fontsize=12)
        ax.set_title('Schedule Performance Index (SPI)', fontsize=14, fontweight='bold')
        ax.axhline(y=1.0, color='black', linestyle='--', linewidth=1, label='Target (1.0)')
        ax.set_ylim([0, max(1.5, comparison.spi * 1.2)])
        ax.legend()
        ax.grid(True, alpha=0.3, axis='y')
        
        # Add value labels on bars
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{height:.3f}',
                   ha='center', va='bottom', fontweight='bold')
        
        chart_path = 'temp_spi_chart.png'
        plt.tight_layout()
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_path
    
    def generate_gantt_chart(self, schedule: Schedule, output_path: str,
                           critical_only: bool = False) -> None:
        """
        Generate Gantt chart for schedule.
        
        Args:
            schedule: Schedule to visualize
            output_path: Output file path
            critical_only: Show only critical path activities
        """
        logger.info(f"Generating Gantt chart: {output_path}")
        
        activities = schedule.get_critical_path() if critical_only else list(schedule.activities.values())
        activities = sorted(activities, key=lambda a: a.start_date if a.start_date else datetime.max)[:50]
        
        fig, ax = plt.subplots(figsize=(14, max(8, len(activities) * 0.3)))
        
        for i, activity in enumerate(activities):
            if not activity.start_date or not activity.finish_date:
                continue
            
            start = mdates.date2num(activity.start_date)
            duration = (activity.finish_date - activity.start_date).days
            
            color = '#d62728' if activity.is_critical else '#1f4788'
            
            ax.barh(i, duration, left=start, height=0.8, 
                   color=color, alpha=0.7, edgecolor='black', linewidth=0.5)
            
            # Add activity name
            ax.text(start - 5, i, activity.activity_name[:40], 
                   va='center', ha='right', fontsize=8)
        
        ax.set_yticks(range(len(activities)))
        ax.set_yticklabels([a.activity_id for a in activities], fontsize=7)
        ax.set_xlabel('Date', fontsize=12)
        ax.set_title(f'Gantt Chart - {schedule.project_name}', fontsize=14, fontweight='bold')
        
        ax.xaxis_date()
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        plt.xticks(rotation=45, ha='right')
        ax.grid(True, alpha=0.3, axis='x')
        
        # Add legend
        from matplotlib.patches import Patch
        legend_elements = [
            Patch(facecolor='#d62728', alpha=0.7, label='Critical'),
            Patch(facecolor='#1f4788', alpha=0.7, label='Non-Critical')
        ]
        ax.legend(handles=legend_elements, loc='upper right')
        
        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        logger.info("Gantt chart generated")
    
    def export_to_excel(self, comparison: ComparisonResult,
                       windows: List[ForensicWindow],
                       output_path: str) -> None:
        """
        Export comprehensive analysis to Excel.
        
        Args:
            comparison: ComparisonResult from analysis
            windows: List of ForensicWindow objects
            output_path: Output Excel file path
        """
        logger.info(f"Exporting to Excel: {output_path}")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Summary sheet
            self._write_summary_sheet(writer, comparison, windows)
            
            # Delayed activities
            self._write_delayed_activities_sheet(writer, comparison)
            
            # Critical path changes
            self._write_critical_path_sheet(writer, comparison)
            
            # Float changes
            self._write_float_changes_sheet(writer, comparison)
            
            # Milestone delays
            if comparison.milestone_delays:
                self._write_milestone_sheet(writer, comparison)
            
            # Window analysis
            if windows:
                self._write_windows_sheet(writer, windows)
        
        logger.info("Excel export complete")
    
    def _write_summary_sheet(self, writer, comparison: ComparisonResult,
                            windows: List[ForensicWindow]) -> None:
        """Write summary sheet to Excel."""
        summary_data = {
            'Metric': [
                'Project Name',
                'Baseline Start',
                'Baseline Finish',
                'Current Start',
                'Current Finish',
                'Overall Project Delay (days)',
                'Schedule Performance Index (SPI)',
                'Completion Variance (%)',
                'Total Activities (Baseline)',
                'Total Activities (Current)',
                'Delayed Activities',
                'Accelerated Activities',
                'Activities Became Critical',
                'Activities Left Critical Path',
                'Milestone Delays',
                'Forensic Windows Analyzed'
            ],
            'Value': [
                comparison.baseline_schedule.project_name,
                comparison.baseline_schedule.start_date.strftime('%Y-%m-%d'),
                comparison.baseline_schedule.finish_date.strftime('%Y-%m-%d'),
                comparison.current_schedule.start_date.strftime('%Y-%m-%d'),
                comparison.current_schedule.finish_date.strftime('%Y-%m-%d'),
                comparison.overall_delay,
                f"{comparison.spi:.3f}",
                f"{comparison.completion_variance:.2f}",
                len(comparison.baseline_schedule.activities),
                len(comparison.current_schedule.activities),
                len(comparison.delayed_activities),
                len(comparison.accelerated_activities),
                len(comparison.new_critical_activities),
                len(comparison.removed_critical_activities),
                len(comparison.milestone_delays),
                len(windows)
            ]
        }
        
        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _write_delayed_activities_sheet(self, writer,
                                       comparison: ComparisonResult) -> None:
        """Write delayed activities sheet."""
        if not comparison.delayed_activities:
            return
        
        delayed_data = []
        for act in comparison.delayed_activities:
            baseline_act = comparison.baseline_schedule.activities[act.activity_id]
            
            delayed_data.append({
                'Activity ID': act.activity_id,
                'Activity Code': act.activity_code,
                'Activity Name': act.activity_name,
                'WBS': act.wbs,
                'Baseline Start': baseline_act.start_date,
                'Baseline Finish': baseline_act.finish_date,
                'Current Start': act.start_date,
                'Current Finish': act.finish_date,
                'Actual Start': act.actual_start,
                'Actual Finish': act.actual_finish,
                'Delay (days)': self._get_activity_delay_days(act, comparison.baseline_schedule),
                'Baseline Float': baseline_act.total_float,
                'Current Float': act.total_float,
                'Is Critical': act.is_critical,
                'Status': act.status.value,
                'Percent Complete': act.percent_complete
            })
        
        df = pd.DataFrame(delayed_data)
        df.to_excel(writer, sheet_name='Delayed Activities', index=False)
    
    def _write_critical_path_sheet(self, writer,
                                   comparison: ComparisonResult) -> None:
        """Write critical path changes sheet."""
        cp_data = []
        
        for act in comparison.new_critical_activities:
            baseline_act = comparison.baseline_schedule.activities[act.activity_id]
            cp_data.append({
                'Activity ID': act.activity_id,
                'Activity Name': act.activity_name,
                'Change': 'Became Critical',
                'Baseline Float': baseline_act.total_float,
                'Current Float': act.total_float,
                'Float Change': act.total_float - baseline_act.total_float
            })
        
        baseline_critical_ids = {
            act_id for act_id, act in comparison.baseline_schedule.activities.items()
            if getattr(act, 'is_critical', False)
        }
        current_critical_ids = {
            act_id for act_id, act in comparison.current_schedule.activities.items()
            if getattr(act, 'is_critical', False)
        }
        
        removed_critical_ids = baseline_critical_ids - current_critical_ids
        for act_id in removed_critical_ids:
            baseline_act = comparison.baseline_schedule.activities.get(act_id)
            current_act = comparison.current_schedule.activities.get(act_id)
            if baseline_act and current_act:
                cp_data.append({
                    'Activity ID': current_act.activity_id,
                    'Activity Name': current_act.activity_name,
                    'Change': 'No Longer Critical',
                    'Baseline Float': baseline_act.total_float,
                    'Current Float': current_act.total_float,
                    'Float Change': current_act.total_float - baseline_act.total_float
                })
        
        if cp_data:
            df = pd.DataFrame(cp_data)
            df.to_excel(writer, sheet_name='Critical Path Changes', index=False)
            
            # Format the sheet
            worksheet = writer.sheets['Critical Path Changes']
            for idx, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).apply(len).max(), len(col)) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 50)
            
            # Apply conditional formatting for change type
            change_col = df.columns.get_loc('Change') + 1
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=change_col)
                if cell.value == 'Became Critical':
                    cell.fill = openpyxl.styles.PatternFill(start_color='FFCCCC', 
                                                           end_color='FFCCCC', 
                                                           fill_type='solid')
                elif cell.value == 'No Longer Critical':
                    cell.fill = openpyxl.styles.PatternFill(start_color='CCFFCC', 
                                                           end_color='CCFFCC', 
                                                           fill_type='solid')
    
    def _write_float_changes_sheet(self, writer: pd.ExcelWriter, comparison: ComparisonResult):
        """Write float changes analysis to Excel."""
        float_data = []
        
        # Match activities between schedules
        activity_matches = {}
        for curr_id, curr_act in comparison.current_schedule.activities.items():
            baseline_act = comparison.baseline_schedule.activities.get(curr_id)
            if baseline_act:
                activity_matches[curr_id] = (baseline_act, curr_act)
        
        # Calculate float changes
        for act_id, (baseline_act, current_act) in activity_matches.items():
            baseline_float = baseline_act.total_float
            current_float = current_act.total_float
            float_change = current_float - baseline_float
            
            # Only include activities with significant float changes (>= 1 day)
            if abs(float_change) >= 1.0:
                # Determine delay cause if activity is delayed
                primary_cause = 'Unknown'
                delay_days = 0
                
                for delay_event in getattr(comparison, 'delay_events', []):
                    if delay_event.activity_id == act_id:
                        primary_cause = getattr(delay_event, 'primary_cause', 'Unknown')
                        delay_days = self._get_delay_event_days(delay_event)
                        break
                
                baseline_duration = self._get_activity_duration(baseline_act)
                current_duration = self._get_activity_duration(current_act)
                duration_change = (current_duration or 0) - (baseline_duration or 0)
                
                float_data.append({
                    'Activity ID': current_act.activity_id,
                    'Activity Name': current_act.activity_name,
                    'Baseline Start': baseline_act.start_date.strftime('%Y-%m-%d') if baseline_act.start_date else '',
                    'Current Start': current_act.start_date.strftime('%Y-%m-%d') if current_act.start_date else '',
                    'Baseline Finish': baseline_act.finish_date.strftime('%Y-%m-%d') if baseline_act.finish_date else '',
                    'Current Finish': current_act.finish_date.strftime('%Y-%m-%d') if current_act.finish_date else '',
                    'Baseline Float (days)': round(baseline_float, 2),
                    'Current Float (days)': round(current_float, 2),
                    'Float Change (days)': round(float_change, 2),
                    'Duration Change (days)': round(duration_change, 2),
                    'Delay Days': round(delay_days, 2),
                    'Primary Cause': primary_cause,
                    'Status': 'Critical' if current_float <= 0 else 'Non-Critical'
                })
        
        # Sort by absolute float change (descending)
        float_data.sort(key=lambda x: abs(x['Float Change (days)']), reverse=True)
        
        if float_data:
            df = pd.DataFrame(float_data)
            df.to_excel(writer, sheet_name='Float Changes', index=False)
            
            # Format the sheet
            worksheet = writer.sheets['Float Changes']
            for idx, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).apply(len).max(), len(col)) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 40)
            
            # Apply conditional formatting for float changes
            float_change_col = df.columns.get_loc('Float Change (days)') + 1
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=float_change_col)
                if cell.value is not None:
                    if cell.value < 0:  # Float decreased (bad)
                        cell.fill = openpyxl.styles.PatternFill(start_color='FFCCCC', 
                                                               end_color='FFCCCC', 
                                                               fill_type='solid')
                    elif cell.value > 0:  # Float increased (good)
                        cell.fill = openpyxl.styles.PatternFill(start_color='CCFFCC', 
                                                               end_color='CCFFCC', 
                                                               fill_type='solid')
            
            # Highlight critical activities
            status_col = df.columns.get_loc('Status') + 1
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=status_col)
                if cell.value == 'Critical':
                    cell.fill = openpyxl.styles.PatternFill(start_color='FFD966', 
                                                           end_color='FFD966', 
                                                           fill_type='solid')
    
    def _write_milestone_sheet(self, writer: pd.ExcelWriter, comparison: ComparisonResult):
        """Write milestone analysis to Excel."""
        milestone_data = []
        
        # Identify milestones (activities with zero duration or milestone flag)
        baseline_milestones = {}
        current_milestones = {}
        
        for act_id, act in comparison.baseline_schedule.activities.items():
            duration = self._get_activity_duration(act) or 0
            if duration == 0 or (hasattr(act, 'is_milestone') and act.is_milestone):
                baseline_milestones[act_id] = act
        
        for act_id, act in comparison.current_schedule.activities.items():
            duration = self._get_activity_duration(act) or 0
            if duration == 0 or (hasattr(act, 'is_milestone') and act.is_milestone):
                current_milestones[act_id] = act
        
        # Compare milestones
        for act_id in baseline_milestones:
            if act_id in current_milestones:
                baseline_ms = baseline_milestones[act_id]
                current_ms = current_milestones[act_id]
                
                # Calculate delay in days
                delay_days = 0
                if baseline_ms.finish_date and current_ms.finish_date:
                    delay_days = (current_ms.finish_date - baseline_ms.finish_date).days
                
                # Determine delay cause
                primary_cause = 'Unknown'
                contributing_causes = []
                
                for delay_event in getattr(comparison, 'delay_events', []):
                    if delay_event.activity_id == act_id:
                        primary_cause = getattr(delay_event, 'primary_cause', 'Unknown')
                        if hasattr(delay_event, 'contributing_causes'):
                            contributing_causes = delay_event.contributing_causes
                        break
                
                milestone_data.append({
                    'Milestone ID': current_ms.activity_id,
                    'Milestone Name': current_ms.activity_name,
                    'Baseline Date': baseline_ms.finish_date.strftime('%Y-%m-%d') if baseline_ms.finish_date else '',
                    'Current Forecast Date': current_ms.finish_date.strftime('%Y-%m-%d') if current_ms.finish_date else '',
                    'Actual Date': current_ms.actual_finish.strftime('%Y-%m-%d') if hasattr(current_ms, 'actual_finish') and current_ms.actual_finish else '',
                    'Delay (days)': round(delay_days, 2),
                    'Status': 'Delayed' if delay_days > 0 else ('On Track' if delay_days == 0 else 'Ahead'),
                    'Primary Cause': primary_cause,
                    'Contributing Causes': ', '.join(contributing_causes) if contributing_causes else '',
                    'Critical': 'Yes' if current_ms.total_float <= 0 else 'No'
                })
        
        # Sort by delay (descending)
        milestone_data.sort(key=lambda x: x['Delay (days)'], reverse=True)
        
        if milestone_data:
            df = pd.DataFrame(milestone_data)
            df.to_excel(writer, sheet_name='Milestone Analysis', index=False)
            
            # Format the sheet
            worksheet = writer.sheets['Milestone Analysis']
            for idx, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).apply(len).max(), len(col)) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 40)
            
            # Apply conditional formatting for status
            status_col = df.columns.get_loc('Status') + 1
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=status_col)
                if cell.value == 'Delayed':
                    cell.fill = openpyxl.styles.PatternFill(start_color='FFCCCC', 
                                                           end_color='FFCCCC', 
                                                           fill_type='solid')
                elif cell.value == 'Ahead':
                    cell.fill = openpyxl.styles.PatternFill(start_color='CCFFCC', 
                                                           end_color='CCFFCC', 
                                                           fill_type='solid')
                elif cell.value == 'On Track':
                    cell.fill = openpyxl.styles.PatternFill(start_color='FFFFCC', 
                                                           end_color='FFFFCC', 
                                                           fill_type='solid')
            
            # Highlight critical milestones
            critical_col = df.columns.get_loc('Critical') + 1
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=critical_col)
                if cell.value == 'Yes':
                    cell.font = openpyxl.styles.Font(bold=True, color='FF0000')
    
    def _write_windows_sheet(self, writer: pd.ExcelWriter, windows: List[ForensicWindow]):
        """Write forensic windows analysis to Excel."""
        if not windows:
            return
        
        window_summary = []
        
        for window in windows:
            # Calculate delay attribution by cause
            cause_delays = {}
            total_window_delay = 0
            
            for delay_event in window.delay_events:
                delay_days = self._get_delay_event_days(delay_event)
                total_window_delay += delay_days
                
                cause = getattr(delay_event, 'primary_cause', 'Unknown')
                if cause not in cause_delays:
                    cause_delays[cause] = 0
                cause_delays[cause] += delay_days
            
            # Create summary row for this window
            window_data = {
                'Window ID': window.window_id,
                'Window Name': getattr(window, 'name', window.window_id),
                'Start Date': window.start_date.strftime('%Y-%m-%d'),
                'End Date': window.end_date.strftime('%Y-%m-%d'),
                'Duration (days)': (window.end_date - window.start_date).days,
                'Total Delay (days)': round(total_window_delay, 2),
                'Number of Delays': len(window.delay_events),
                'Critical Path Impact': round(window.critical_path_delay, 2) if hasattr(window, 'critical_path_delay') else 0,
                'Schedule Performance Index': round(window.spi, 3) if hasattr(window, 'spi') else 1.0
            }
            
            # Add delay by cause columns
            for cause, delay_days in sorted(cause_delays.items(), key=lambda x: x[1], reverse=True):
                window_data[f'{cause} (days)'] = round(delay_days, 2)
            
            window_summary.append(window_data)
        
        if window_summary:
            df = pd.DataFrame(window_summary)
            df = df.fillna(0)  # Fill NaN values with 0 for missing causes
            df.to_excel(writer, sheet_name='Forensic Windows', index=False)
            
            # Format the sheet
            worksheet = writer.sheets['Forensic Windows']
            for idx, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).apply(len).max(), len(col)) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 35)
            
            # Apply conditional formatting for SPI
            if 'Schedule Performance Index' in df.columns:
                spi_col = df.columns.get_loc('Schedule Performance Index') + 1
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=spi_col)
                    if cell.value is not None:
                        if cell.value < 0.9:  # Behind schedule
                            cell.fill = openpyxl.styles.PatternFill(start_color='FFCCCC', 
                                                                   end_color='FFCCCC', 
                                                                   fill_type='solid')
                        elif cell.value >= 1.0:  # On or ahead of schedule
                            cell.fill = openpyxl.styles.PatternFill(start_color='CCFFCC', 
                                                                   end_color='CCFFCC', 
                                                                   fill_type='solid')
                        else:  # Slightly behind (0.9-1.0)
                            cell.fill = openpyxl.styles.PatternFill(start_color='FFFFCC', 
                                                                   end_color='FFFFCC', 
                                                                   fill_type='solid')
        
        # Create detailed window sheet with individual delay events
        self._write_window_details_sheet(writer, windows)
    
    def _write_window_details_sheet(self, writer: pd.ExcelWriter, windows: List[ForensicWindow]):
        """Write detailed delay events for each forensic window."""
        if not windows:
            return
        
        detail_data = []
        
        for window in windows:
            for delay_event in window.delay_events:
                delay_days = self._get_delay_event_days(delay_event)
                
                detail_data.append({
                    'Window ID': window.window_id,
                    'Window Name': getattr(window, 'name', window.window_id),
                    'Activity ID': delay_event.activity_id,
                    'Activity Name': delay_event.activity_name,
                    'Delay Type': delay_event.delay_type.value if isinstance(delay_event.delay_type, Enum) else str(delay_event.delay_type),
                    'Primary Cause': getattr(delay_event, 'primary_cause', 'Unknown'),
                    'Delay Days': round(delay_days, 2),
                    'Is Critical': 'Yes' if delay_event.is_critical else 'No',
                    'Is Concurrent': 'Yes' if delay_event.is_concurrent else 'No',
                    'Baseline Start': delay_event.baseline_start.strftime('%Y-%m-%d') if delay_event.baseline_start else '',
                    'Actual Start': delay_event.actual_start.strftime('%Y-%m-%d') if delay_event.actual_start else '',
                    'Baseline Finish': delay_event.baseline_finish.strftime('%Y-%m-%d') if delay_event.baseline_finish else '',
                    'Actual Finish': delay_event.actual_finish.strftime('%Y-%m-%d') if delay_event.actual_finish else '',
                    'Description': delay_event.description if hasattr(delay_event, 'description') else ''
                })
        
        if detail_data:
            df = pd.DataFrame(detail_data)
            df.to_excel(writer, sheet_name='Window Details', index=False)
            
            # Format the sheet
            worksheet = writer.sheets['Window Details']
            for idx, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).apply(len).max(), len(col)) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 35)
            
            # Apply conditional formatting for critical delays
            critical_col = df.columns.get_loc('Is Critical') + 1
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=critical_col)
                if cell.value == 'Yes':
                    # Highlight entire row for critical delays
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = openpyxl.styles.PatternFill(
                            start_color='FFE6E6', 
                            end_color='FFE6E6', 
                            fill_type='solid'
                        )
    
    def _get_delay_event_days(self, delay_event: DelayEvent) -> float:
        """Calculate total delay days for a delay event."""
        delay_days = 0.0
        
        if delay_event.actual_start and delay_event.baseline_start:
            start_delay = (delay_event.actual_start - delay_event.baseline_start).days
            if start_delay > 0:
                delay_days += start_delay
        
        if delay_event.actual_finish and delay_event.baseline_finish:
            finish_delay = (delay_event.actual_finish - delay_event.baseline_finish).days
            if finish_delay > 0:
                delay_days += finish_delay
        
        # If duration information is available
        if hasattr(delay_event, 'baseline_duration') and hasattr(delay_event, 'actual_duration'):
            duration_delay = delay_event.actual_duration - delay_event.baseline_duration
            if duration_delay > 0:
                delay_days = max(delay_days, duration_delay)
        
        return delay_days